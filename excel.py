import openpyxl
import re
import os
from openpyxl import Workbook as ws
from openpyxl import load_workbook

# Получаем текущую директорию, где находится скрипт
script_directory = os.path.dirname(os.path.abspath(__file__))

# Указываем полный путь к файлу wildberries.xlsx
wildberries_path = os.path.join(script_directory, 'wildberries.xlsx')

# Проверяем, существует ли файл
if not os.path.isfile(wildberries_path):
    raise FileNotFoundError(f"Файл {wildberries_path} не найден!")

# Загружаем файл
workbook = openpyxl.load_workbook(wildberries_path)

# Выбираю лист который мне нужен
sheet = workbook['Лист подбора']

# Список для хранения данных из столбца
column_l = []
M = []

# Проходим по всем строкам из столбца G
for cell in sheet["G"]:
    # Пропускаем заголовок
    if cell.value == "Артикул продавца":
        continue
    # Нормализация артикула (приведение к нижнему регистру и удаление пробелов)
    articul_normalized = str(cell.value).strip().lower()
    column_l.append(articul_normalized)

# Функция для удаления знаков из названия файла
def remove_special_chars(name):
    return re.sub('[\\/:*?"<>|]', '', name)

for file in column_l:
    if file is not None:
        file_name = remove_special_chars(file)
        M.append(file_name + '.rtf')

# Указываем полный путь к файлу products.xlsx
products_path = os.path.join(script_directory, 'products.xlsx')

# Проверяем, существует ли файл
if not os.path.isfile(products_path):
    raise FileNotFoundError(f"Файл {products_path} не найден!")

# Загружаем данные из products.xlsx
products_workbook = openpyxl.load_workbook(products_path)
products_sheet = products_workbook['Лист1']  # Убедитесь, что имя листа правильное

# Создаем словарь для хранения артикулов и стеллажей
products_dict = {}

# Проходим по всем строкам в products.xlsx
for row in products_sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
    articul = row[1]  # Артикул находится в столбце B (индекс 1)
    shelf = row[2]    # Стеллаж находится в столбце C (индекс 2)
    if articul is not None:  # Пропускаем пустые строки
        # Нормализация артикула перед добавлением в словарь
        articul_normalized = str(articul).strip().lower()
        # Удаляем символ '/' из артикула
        articul_normalized = articul_normalized.replace('/', '')
        products_dict[articul_normalized] = shelf



# Функция для получения стеллажа по артикулу
def get_shelf(articul):
    # Нормализация артикула (удаление пробелов и приведение к нижнему регистру)
    articul_normalized = str(articul).strip().lower()
    # Удаляем символ '/' из артикула
    articul_normalized = articul_normalized.replace('/', '')
    return products_dict.get(articul_normalized, "None")